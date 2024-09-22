import datetime
import openpyxl
from typing import List
from io import BytesIO
from app.model.debit_notes import DebitNoteCWUpload

def create_cw_upload_file(request:List[DebitNoteCWUpload]):
    file_path = 'app/assets/CW1_NewChargeRate_UpSert.xlsx'
    
    # Load the existing workbook and get the active worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Start appending from cell A2
    start_row = 2
    # Remove any existing data starting from row 2
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Append the data from the request to the worksheet
    start_row = 2
    current_row = start_row

    for row in request:
        ws.cell(row=current_row, column=1, value=row.agl_shipment_number)
        ws.cell(row=current_row, column=2, value =row.creditor)
        ws.cell(row=current_row, column=3, value =row.invoice_num)
        ws.cell(row=current_row, column=4, value =row.invoice_date)
        ws.cell(row=current_row, column=5, value =row.supplier_cost_ref)
        ws.cell(row=current_row, column=6, value =row.ar_ap)
        ws.cell(row=current_row, column=7, value =row.is_post)
        ws.cell(row=current_row, column=8, value =row.FRT)
        ws.cell(row=current_row, column=9, value =row.AGEN)
        ws.cell(row=current_row, column=10, value =row.AMS)
        ws.cell(row=current_row, column=11, value =row.EQUIP)
        ws.cell(row=current_row, column=12, value =row.ORIGIN)
        ws.cell(row=current_row, column=13, value =row.LOCAL)
        ws.cell(row=current_row, column=14, value =row.PP)
        ws.cell(row=current_row, column=15, value =row.TERFEE)
        ws.cell(row=current_row, column=16, value =row.TELEX)
        ws.cell(row=current_row, column=17, value =row.VGM)
        ws.cell(row=current_row, column=18, value =row.DCART)
        ws.cell(row=current_row, column=19, value =row.SEAL)
        ws.cell(row=current_row, column=20, value =row.DOC)
        ws.cell(row=current_row, column=21, value =row.BOOK)
        ws.cell(row=current_row, column=22, value =row.CCLR)
        ws.cell(row=current_row, column=23, value =row.ALINE)
        ws.cell(row=current_row, column=24, value =row.PSEC)
        ws.cell(row=current_row, column=25, value =row.EGF)
        ws.cell(row=current_row, column=26, value =row.TRANS)
        ws.cell(row=current_row, column=27, value =row.CHRENT)
        current_row += 1
        
    # Save the workbook to a BytesIO object
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream